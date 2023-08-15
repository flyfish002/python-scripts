import openpyxl
import subprocess
import sys

if len(sys.argv) < 3:
   print("Please provide at least two parameters:  xlsx name, sheet name") 
   sys.exit(1)

xlsx_name  = sys.argv[1] + ".xlsx"
sheet_name = sys.argv[2] 
yaml_name  = sys.argv[1] + "-" + sheet_name + ".yaml" 

# xlsx_name  = "20230811-AndyUSIPs.xlsx"
# sheet_name = "IP1"
# yaml_name  = "20230811-AndyUSIPs-IP1.yaml"

# print("xlsx_name:",  xlsx_name)
# print("sheet_name:", sheet_name) 
# print("yaml_name:", yaml_name)
# sys.exit(1)

# 打开Excel文件
workbook = openpyxl.load_workbook(xlsx_name)

# 选择工作表（Sheet）
sheet = workbook[sheet_name]

# 获取最大行数和列数
max_row = sheet.max_row
max_column = sheet.max_column

# 清空yaml文件
empty_content = ""
clear_yaml_command = 'echo "{}" > "{}"'.format(empty_content,yaml_name)
subprocess.run(clear_yaml_command, shell=True)

# 开头填入 services
begin_content = "services:"
write_begin_command = 'echo "{}" > "{}"'.format(begin_content,yaml_name)  
subprocess.run(write_begin_command, shell=True)

# 遍历每一行
for row in range(2, max_row+1):  # 从第二行开始，跳过第一行
    # 跳过空行
    if not sheet.cell(row=row, column=1).value:
#        print("存在空行，需跳过！！！！！！！！！！！！！！！！！！！") 
       continue
    # 赋值每行数据
    proxy_ip = sheet.cell(row=row, column=1).value
    proxy_port = sheet.cell(row=row, column=2).value
    proxy_handler_protocol = sheet.cell(row=row, column=3).value 
    proxy_listener_protocol = sheet.cell(row=row, column=4).value
    proxy_username = sheet.cell(row=row, column=5).value
    proxy_userpasswd = sheet.cell(row=row, column=6).value

#    print("              ")
#    print("proxy_ip: %s" % proxy_ip)    
#    print("proxy_port: %s" % proxy_port)
#    print("proxy_handler_protocol: %s" % proxy_handler_protocol)
#    print("proxy_listener_protocol: %s" % proxy_listener_protocol)
#    print("proxy_username: %s" % proxy_username)
#    print("proxy_userpasswd: %s" % proxy_userpasswd)
#    print("              ") 

    name_content ="- name: service-" + str(row)
    write_name_command = 'echo "{}" >> "{}"'.format(name_content,yaml_name) 
    subprocess.run(write_name_command, shell=True)

    addr_port_content ='  addr: "' + str(proxy_ip) + ':' + str(proxy_port) + '"'    
    write_add_port_command = 'echo "{}" >> "{}"'.format(addr_port_content,yaml_name)
    subprocess.run(write_add_port_command, shell=True)

    interface_content = "  interface: " + str(proxy_ip)   
    write_interface_command = 'echo "{}" >> "{}"'.format(interface_content,yaml_name)
    subprocess.run(write_interface_command, shell=True)

    handler_content = '  handler:'
    handler_type_content = '    type: ' + proxy_handler_protocol  
    write_handler_command1 = 'echo "{}" >> "{}";echo "{}" >> "{}"'.format(handler_content,yaml_name,handler_type_content,yaml_name) 
    subprocess.run(write_handler_command1, shell=True) 
    handler_auth_content = '    auth:'
    handler_auth_user_content = '      username: ' + proxy_username 
    handler_auth_passwd_content = '      password: ' + proxy_userpasswd
    write_handler_command2 = 'echo "{}" >> "{}";echo "{}" >> "{}";echo "{}" >> "{}"'.format(handler_auth_content,yaml_name,handler_auth_user_content,yaml_name,handler_auth_passwd_content,yaml_name)
    subprocess.run(write_handler_command2, shell=True)
    metedata_content = '    metadata:'
    metedata_tls_content = '      notls: ' +  '"true"'
    write_handler_command3 = 'echo "{}" >> "{}";echo "{}" >> "{}"'.format(metedata_content,yaml_name,metedata_tls_content,yaml_name)
    subprocess.run(write_handler_command3, shell=True)

    listener_content = '  listener:'
    listener_type_content = '    type: ' + proxy_listener_protocol
    write_listener_command = 'echo "{}" >> "{}";echo "{}" >> "{}"'.format(listener_content,yaml_name,listener_type_content,yaml_name)
    subprocess.run(write_listener_command, shell=True)

    write_empty_raw_command = 'echo "{}" >> "{}"'.format(empty_content,yaml_name)  
    subprocess.run(write_empty_raw_command, shell=True)

# 关闭Excel文件
workbook.close()
